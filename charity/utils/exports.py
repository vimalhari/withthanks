import csv
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def format_duration(seconds):
    """Format seconds to mm:ss"""
    if seconds is None:
        return "00:00"
    minutes = seconds // 60
    secs = seconds % 60
    return f"{minutes:02d}:{secs:02d}"


def export_analytics_csv(data_list, filename, summary_stats=None):
    output = io.StringIO()
    writer = csv.writer(output)

    # 1. Summary Section (if provided)
    if summary_stats:
        writer.writerow(["SUMMARY REPORT"])
        writer.writerow(["Total Views", summary_stats.get("views", 0)])
        writer.writerow(["Success Rate", f"{summary_stats.get('success_rate', 0)}%"])
        writer.writerow(["Total Processed", summary_stats.get("processed", 0)])
        writer.writerow([])  # Empty line

    # 2. Detailed Headers
    writer.writerow(
        [
            "Date",
            "Campaign",
            "Recipients",
            "Delivered",
            "Opened",
            "Clicked",
            "Plays",
            "Avg Watch Time",
            "Completion %",
        ]
    )

    for row in data_list:
        writer.writerow(
            [
                row["date"],
                row["campaign_name"],
                row["recipients"],
                row["delivered"],
                row["opened"],
                row["clicked"],
                row["plays"],
                format_duration(row["avg_watch_time"]),
                f"{row['completion_rate']}%",
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def export_analytics_excel(data_list, filename, summary_stats=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    centered_alignment = Alignment(horizontal="center")
    bold_font = Font(bold=True)

    current_row = 1

    # 1. Summary Section (if provided)
    if summary_stats:
        ws.cell(row=current_row, column=1, value="SUMMARY REPORT").font = bold_font
        current_row += 1

        summary_items = [
            ("Total Views", summary_stats.get("views", 0)),
            ("Success Rate", f"{summary_stats.get('success_rate', 0)}%"),
            ("Total Processed", summary_stats.get("processed", 0)),
        ]

        for label, value in summary_items:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value).font = bold_font
            current_row += 1

        current_row += 1  # Spacer

    # 2. Detailed Headers
    headers = [
        "Date",
        "Campaign",
        "Recipients",
        "Delivered",
        "Opened",
        "Clicked",
        "Plays",
        "Avg Watch Time",
        "Completion %",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment

    header_row_num = current_row

    # 3. Data Rows
    current_row += 1
    start_data_row = current_row

    for row_data in data_list:
        ws.append(
            [
                row_data["date"],
                row_data["campaign_name"],
                row_data["recipients"],
                row_data["delivered"],
                row_data["opened"],
                row_data["clicked"],
                row_data["plays"],
                format_duration(row_data["avg_watch_time"]),
                row_data["completion_rate"] / 100.0,
            ]
        )

    last_data_row = ws.max_row

    # 4. Add Charts (Activity Trends)
    # 4. Add Charts (Activity Trends - Combo Chart)
    if len(data_list) > 0:
        # 1. Bar Chart for "Recipients" (Sent)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.y_axis.title = "Volume"
        bar_chart.x_axis.title = "Date"

        # Recipients Data (Column C -> 3)
        c_recip = Reference(ws, min_col=3, min_row=header_row_num, max_row=last_data_row)
        bar_chart.add_data(c_recip, titles_from_data=True)

        # Categories (Date)
        cats = Reference(ws, min_col=1, min_row=start_data_row, max_row=last_data_row)
        bar_chart.set_categories(cats)

        # 2. Line Chart for "Opened" and "Plays"
        line_chart = LineChart()
        line_chart.style = 12  # Different style for contrast
        line_chart.y_axis.axId = 200  # Secondary axis logic if needed, but here we overlay

        # Opened (Column E -> 5)
        c_open = Reference(ws, min_col=5, min_row=header_row_num, max_row=last_data_row)
        # Plays (Column G -> 7)
        c_play = Reference(ws, min_col=7, min_row=header_row_num, max_row=last_data_row)

        line_chart.add_data(c_open, titles_from_data=True)
        line_chart.add_data(c_play, titles_from_data=True)

        # Combine: Add Line Chart ~to~ Bar Chart
        bar_chart += line_chart

        bar_chart.title = "Activity Trends (Sent vs Engaged)"
        bar_chart.height = 12
        bar_chart.width = 20

        # Place chart next to data
        ws.add_chart(bar_chart, f"K{header_row_num}")

    # Add Totals Row
    totals_row = last_data_row + 1
    ws.cell(row=totals_row, column=1, value="TOTALS").font = bold_font

    for col in range(3, 8):
        col_letter = get_column_letter(col)
        ws.cell(
            row=totals_row,
            column=col,
            value=f"=SUM({col_letter}{start_data_row}:{col_letter}{last_data_row})",
        ).font = bold_font

    # Formatting
    for row in ws.iter_rows(min_row=start_data_row, max_row=totals_row, min_col=9, max_col=9):
        for cell in row:
            cell.number_format = "0.0%"

    # Auto-adjust column width
    for i, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def export_batch_detail_csv(batch_info, data_list, filename):
    output = io.StringIO()
    writer = csv.writer(output)

    # Batch Info Header
    writer.writerow(["Batch Report", batch_info["name"]])
    writer.writerow(["Date", batch_info["date"]])
    writer.writerow([])

    # Headers
    writer.writerow(
        [
            "Recipient Name",
            "Email",
            "Type",
            "Status",
            "Real Hits",
            "Total Views",
            "Watch Duration (s)",
            "Bounce Reason",
        ]
    )

    for row in data_list:
        writer.writerow(
            [
                row["name"],
                row["email"],
                row["type"],
                row["status"],
                row["real_views"],
                row["total_views"],
                row["duration"],
                row["bounce_reason"] or "",
            ]
        )

    response = HttpResponse(output.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def export_batch_detail_excel(batch_info, data_list, filename, timeline_stats=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Detail Report"

    # Batch Headers
    ws.append(["Batch Report:", batch_info["name"]])
    ws.append(["Date:", batch_info["date"]])
    ws.append([])

    # Calculate Summaries for View Chart
    total_real = sum(r["real_views"] for r in data_list)

    # Calculate Summaries for Status Chart
    status_counts = {"Delivered": 0, "Bounced": 0, "Failed": 0}
    for r in data_list:
        s = r.get("status", "").lower()
        if "bounced" in s:
            status_counts["Bounced"] += 1
        elif "fail" in s:
            status_counts["Failed"] += 1
        elif "success" in s or "delivered" in s:
            status_counts["Delivered"] += 1

    # Hidden Summary Table for Charts
    # View Dist
    summary_row_start = 5
    ws.cell(row=summary_row_start, column=1, value="Metric")
    ws.cell(row=summary_row_start, column=2, value="Count")

    ws.cell(row=summary_row_start + 1, column=1, value="Real Views")
    ws.cell(row=summary_row_start + 1, column=2, value=total_real)

    # Status Dist (Spacing it out)
    status_row_start = summary_row_start + 4
    ws.cell(row=status_row_start, column=1, value="Status")
    ws.cell(row=status_row_start, column=2, value="Count")

    ws.cell(row=status_row_start + 1, column=1, value="Delivered")
    ws.cell(row=status_row_start + 1, column=2, value=status_counts["Delivered"])
    ws.cell(row=status_row_start + 2, column=1, value="Bounced")
    ws.cell(row=status_row_start + 2, column=2, value=status_counts["Bounced"])
    ws.cell(row=status_row_start + 3, column=1, value="Failed")
    ws.cell(row=status_row_start + 3, column=2, value=status_counts["Failed"])

    # Timeline Stats (Hidden columns for Combo Chart)
    combo_chart_ref = None
    if timeline_stats:
        timeline_row_start = status_row_start + 5
        ws.cell(row=timeline_row_start, column=1, value="Date")
        ws.cell(row=timeline_row_start, column=2, value="Sent")
        ws.cell(row=timeline_row_start, column=3, value="Opened")
        ws.cell(row=timeline_row_start, column=4, value="Played")

        for i, stat in enumerate(timeline_stats):
            r = timeline_row_start + 1 + i
            ws.cell(row=r, column=1, value=stat["date"])
            ws.cell(row=r, column=2, value=stat["sent"])
            ws.cell(row=r, column=3, value=stat["opened"])
            ws.cell(row=r, column=4, value=stat["played"])

        last_timeline_row = timeline_row_start + len(timeline_stats)

        # Build Combo Chart
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.y_axis.title = "Volume"
        bar_chart.x_axis.title = "Date"

        # Sent (Col 2)
        c_sent = Reference(ws, min_col=2, min_row=timeline_row_start, max_row=last_timeline_row)
        bar_chart.add_data(c_sent, titles_from_data=True)

        cats = Reference(ws, min_col=1, min_row=timeline_row_start + 1, max_row=last_timeline_row)
        bar_chart.set_categories(cats)

        # Line Chart for Opened/Played
        line_chart = LineChart()
        line_chart.style = 12
        line_chart.y_axis.axId = 200

        c_open = Reference(ws, min_col=3, min_row=timeline_row_start, max_row=last_timeline_row)
        c_play = Reference(ws, min_col=4, min_row=timeline_row_start, max_row=last_timeline_row)

        line_chart.add_data(c_open, titles_from_data=True)
        line_chart.add_data(c_play, titles_from_data=True)

        bar_chart += line_chart
        bar_chart.title = "Batch Activity Trends"
        bar_chart.height = 12
        bar_chart.width = 24

        combo_chart_ref = bar_chart

    ws.append([])  # Spacer

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

    headers = [
        "Recipient Name",
        "Email",
        "Type",
        "Status",
        "Real Hits",
        "Total Views",
        "Watch Duration (s)",
        "Bounce Reason",
    ]
    ws.append(headers)

    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill

    for row in data_list:
        ws.append(
            [
                row["name"],
                row["email"],
                row["type"],
                row["status"],
                row["real_views"],
                row["total_views"],
                row["duration"],
                row["bounce_reason"] or "",
            ]
        )

    # 1. View Sources Pie Chart
    pie_views = PieChart()
    pie_views.title = "View Sources"
    label_v = Reference(ws, min_col=1, min_row=summary_row_start + 1, max_row=summary_row_start + 1)
    data_v = Reference(ws, min_col=2, min_row=summary_row_start + 1, max_row=summary_row_start + 1)
    pie_views.add_data(data_v, titles_from_data=False)
    pie_views.set_categories(label_v)
    ws.add_chart(pie_views, "K4")

    # 2. Delivery Status Pie Chart
    pie_status = PieChart()
    pie_status.title = "Delivery Status"
    label_s = Reference(ws, min_col=1, min_row=status_row_start + 1, max_row=status_row_start + 3)
    data_s = Reference(ws, min_col=2, min_row=status_row_start + 1, max_row=status_row_start + 3)
    pie_status.add_data(data_s, titles_from_data=False)
    pie_status.set_categories(label_s)
    ws.add_chart(pie_status, "S4")

    # 3. Combo Chart (if data exists)
    if combo_chart_ref:
        ws.add_chart(combo_chart_ref, "K22")

    # Column widths
    for i, _ in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
